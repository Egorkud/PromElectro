from bs4 import BeautifulSoup
import requests
import time
import random
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains

from instruments.BaseParser import BaseParser
from instruments.BaseParser import ParserLogger

# It would be possibly better to use Selenium everywhere here
class GrandInstrument(BaseParser):
    def __init__(self):
        super().__init__()
        self.driver = webdriver.Chrome()

    def scrap(self, filename_new: str = "scrapped_data(chars) last data.xlsx",
              data_file: str = "unique_finale_with_spaces(no jtc) — копия.xlsx"):

        work_sheet = openpyxl.open(data_file).active
        logger = ParserLogger()

        # Row for new file, count for old file: collecting articules
        for row in range(2, work_sheet.max_row + 1):
            print(self.GREEN(f"{row}. Started"))

            item_articule = work_sheet.cell(row, 1).value
            item_articule = item_articule.replace(" ", "-")

            url_search_ru = "https://grandinstrument.ua/"    # Need to add searching link
            url = f"{url_search_ru}{item_articule}"

            searched_item_link = self.get_searched_item_link(url)
            print(searched_item_link)

            self.log_data = [row - 1 ,item_articule, searched_item_link]

            if not searched_item_link:
                print(self.GREEN(f"Link skipped. None response"))
                logger.log_parsing_result(self.log_data)
                row += 1
                time.sleep(1 + random.uniform(0, 1))
                continue

            self.blank_sheet.cell(row, 1).value = item_articule
            self.blank_sheet.cell(row, 2).value = item_articule

            # If there is necessary to scrap UKR version, it is important for names and descriptions
            ukr_link = searched_item_link.replace("grandinstrument.ua/", "grandinstrument.ua/ua/")  # Old link to new UKR link
            two_lang_links = (searched_item_link, ukr_link)

            req = requests.get(searched_item_link, headers=self.headers)
            soup = BeautifulSoup(req.text, "lxml")

            """Обираємо необхідний парсер і реалізуємо його роботу в окремому файлі"""
            # Uncomment necessary scrappers
            self.get_characteristics(soup, row, searched_item_link)
            # self.download_instruction(soup, row)
            # self.get_descriptions(two_lang_links, row)
            # self.get_photos(soup, row)
            # self.get_product_name(two_lang_links, row)

            logger.log_parsing_result(self.log_data)

            self.blank_file.save(filename_new)
            time.sleep(1 + random.uniform(1, 3))

            row += 1
            row += 1
        self.driver.quit()
        print(self.GREEN(f"\nFile {filename_new} created"))
        print(self.GREEN(f"Total photo count: {self.images_counter}"))
        print(self.GREEN(f"Total descriptions count: {self.instructions_counter}"))

    def get_searched_item_link(self, url):
        """Шукає товар за артикулом та повертає його URL"""
        try:
            req = requests.get(url, headers=self.headers)
            soup = BeautifulSoup(req.text, "lxml")
            searched_item_link = soup.find("h1", class_="page-heading").text.strip()  # Реалізація пошуку
            if searched_item_link == "Страница не найдена (404)":
                return None
            return url

        except Exception as ex:
            self.log_data.append(f"Error getting search link: {ex}")
            return None

    def get_characteristics(self, soup, row, searched_item_link):
        """Отримує характеристики товару"""
        self.driver.get(searched_item_link)
        try:
            try:
                characteristics_vendor = soup.find("div", class_="col-12 product-info-vendor").find_all("ul")
                for ul in characteristics_vendor:
                    all_li = ul.find_all("li")
                    for li in all_li:
                        try:
                            key, value = li.text.split(":")
                            # Check if new char name is not in char dick
                            self.check_key(key)

                            # Adding characteristics
                            char_col = self.char_dict[key]
                            self.blank_sheet.cell(row, char_col).value = value
                        except:
                            pass
            except:
                pass

            try:
                button_language = self.driver.find_element(By.CLASS_NAME, "header-locale-switcher ")
                ActionChains(self.driver).move_to_element(button_language).perform()
                button_language.click()
                time.sleep(2)

                # Перевіряємо, чи є кнопка "Показать все"
                button_chars = self.driver.find_element(By.CLASS_NAME, "chars-show")

                # Прокручуємо сторінку до кнопки (якщо потрібно)
                ActionChains(self.driver).move_to_element(button_chars).perform()

                # Натискаємо на кнопку
                button_chars.click()
                time.sleep(2)  # Чекаємо, поки характеристики завантажаться
            except:
                print("Кнопка 'Показать все' відсутня. Продовжуємо парсинг.")

            # Парсимо характеристики
            characteristics = self.driver.find_element(By.CLASS_NAME, "table-properties")
            char_lines = characteristics.find_elements(By.CLASS_NAME, "row")

            for char in char_lines:
                try:
                    key = char.find_element(By.CLASS_NAME, "property-name").text.strip()
                    value = char.find_element(By.CLASS_NAME, "property-value").text.strip().replace("\n", " ")

                    self.check_key(key)

                    # Adding characteristics
                    char_col = self.char_dict[key]
                    self.blank_sheet.cell(row, char_col).value = value
                except:
                    pass

        except Exception as ex:
            self.log_data.append(f"Error getting characteristics: {ex}")


    def download_instruction(self, soup, row):
        # Get instructions
        try:
            # Implement data scrapper
            find_instr = soup.find(...)

            instruction_link = ...

            self.download_instruction_file(instruction_link, row)

        except Exception as ex:
            self.log_data.append(f"Error getting instructions {ex}")

    def get_descriptions(self, two_lang_links, row):
        # Get description RU (But now it is better to generate with gpt)
        # OPTIONAL: Get description UKR if it is possible on website (other option is translator)
        for idx, language_link in enumerate(two_lang_links):
            try:
                req = requests.get(language_link, headers=self.headers)
                soup_descr = BeautifulSoup(req.text, "lxml")

                description_lines = soup_descr.find("div", itemprop="description").contents
                clean_descriptions = [str(i) for i in description_lines if i != "\n"]

                self.blank_sheet.cell(row, 11 + idx).value = "\n".join(clean_descriptions)
                time.sleep(1 + random.uniform(1, 2))
            except Exception as ex:
                self.log_data.append(f"Error getting descriptions: {ex}")

    def get_photos(self, soup, row):
        # Get photos
        try:
            # Implement scrapping
            photo_data = (soup.find("div", class_="swiper-wrapper")
                          .find_all("a"))

            photo_links = [link.get("href") for link in photo_data]

            self.download_photos(photo_links, row, "folder_name") # Input folder name
        except Exception as ex:
            self.log_data.append(f"Error getting photos: {ex}")

    def get_product_name(self, two_lang_links, row):
        # Get names ru ukr
        for idx, language_link in enumerate(two_lang_links):
            try:
                req = requests.get(language_link, headers=self.headers)
                soup_name = BeautifulSoup(req.text, "lxml")

                # Implement scrapping
                full_name = soup_name.find("h1", class_="page-heading").text

                self.blank_sheet.cell(row, 4 + idx).value = full_name

                # # Save names data
                # self.save_names_data("filename", item_type, last_name, item_articule, series, manufacturer, row, idx)

            except Exception as ex:
                self.log_data.append(f"Error getting product name {ex}")