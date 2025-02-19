from bs4 import BeautifulSoup
import os
import requests
import time
import random
from pathlib2 import Path
from urllib.parse import urlparse
import openpyxl

from instruments.BaseParser import BaseParser

# Only for copy and make new data scrapper for new website
class Noveen_com(BaseParser):
    def __init__(self):
        super().__init__()


    def scrap(self, filename: str = "New_noveen_parsed.xlsx",
              export_file: str = "noveen_parse.xlsx"):

        work_sheet = openpyxl.open(export_file).active

        for row in range(2, work_sheet.max_row + 1):
            print(self.GREEN(f"{row}. Started"))

            item_articule = work_sheet.cell(row, 2).value
            self.blank_sheet.cell(row, 2).value = item_articule

            url_search_ru = "website searching ling"    # Need to add searching link
            url = f"{url_search_ru}{item_articule}"

            # Here we have already all the links
            searched_item_link = work_sheet.cell(row, 3).value
            print(searched_item_link)

            if not searched_item_link:
                print(self.GREEN(f"Link skipped. None response"))
                continue

            # If there is necessary to scrap UKR version, it is important for names and descriptions
            ru_link = searched_item_link.replace("noveen.com.ua/", "noveen.com.ua/ru/")  # Old link to new UKR link
            two_lang_links = (ru_link, searched_item_link)

            req = requests.get(searched_item_link, headers=self.headers)
            soup = BeautifulSoup(req.text, "lxml")

            """Обираємо необхідний парсер і реалізуємо його роботу в окремому файлі"""
            # Uncomment necessary scrappers
            self.get_characteristics(soup, row)
            # self.download_instruction(soup, row)
            self.get_descriptions(two_lang_links, row)
            self.get_photos(soup, row)
            self.get_product_name(two_lang_links, row, item_articule, work_sheet)


            self.blank_file.save(filename)
            time.sleep(1 + random.uniform(1, 3))
        print(self.GREEN(f"\nFile {filename} created"))
        print(self.GREEN(f"Total photo count: {self.images_counter}"))
        print(self.GREEN(f"Total descriptions count: {self.instructions_counter}"))


    def get_searched_item_link(self, url):
        """Шукає товар за артикулом та повертає його URL"""
        try:
            req = requests.get(url, headers=self.headers)
            soup = BeautifulSoup(req.text, "lxml")
            searched_item_link = soup.find(...)  # Реалізація пошуку
            return searched_item_link
        except Exception as ex:
            print(self.RED(f"Error getting search link: {ex}"))
            return None

    def get_characteristics(self, soup,  row):
        """Отримує характеристики товару"""
        try:
            characteristics = (soup.find("div", class_="changePropertiesGroup")
                               .find_all("tr"))  # Get characteristics

            # Get keys values (can be some for find_all loops)
            # Implement scrap
            for tr in characteristics:
                tds = tr.find_all("td", limit=2)

                key, value = tds[0].text, tds[1].text

                # Check if new char name is not in char dick
                self.check_key(key)

                # Adding characteristics
                char_col = self.char_dict[key]
                self.blank_sheet.cell(row, char_col).value = value

        except Exception as ex:
            print(ex)
            print("No characteristics")

    def download_instruction(self, soup, row):
        # Get instructions
        try:
            # Implement data scrapper
            find_instr = soup.find(...)

            instruction_link = ...

            self.download_instruction_file(instruction_link, row)

        except Exception as ex:
            print(ex)
            print("No instructions")

    def get_descriptions(self, two_lang_links, row):
        # Get description RU (But now it is better to generate with gpt)
        # OPTIONAL: Get description UKR if it is possible on website (other option is translator)
        for idx, language_link in enumerate(two_lang_links):
            try:
                req = requests.get(language_link, headers=self.headers)
                soup_descr = BeautifulSoup(req.text, "lxml")

                description_lines = (soup_descr.find("div", id="detailText")
                                    .find("div", class_="heading")
                                    .find_next())
                clean_descriptions = [str(i) for i in description_lines if i != "\n"]

                self.blank_sheet.cell(row, 11 + idx).value = "\n".join(clean_descriptions)
                time.sleep(1 + random.uniform(1, 2))
            except Exception as ex:
                print(ex)
                print("No description")

    def get_photos(self, soup, row):
        # Get photos
        try:
            # Implement scrapping
            photo_data = (soup.find("div", class_="pictureSlider")
                          .find_all("a"))

            photo_links = [link.get("href") for link in photo_data]

            self.download_photos(photo_links, row, "noveen")
        except Exception as ex:
            print(ex)
            print("No photos")

    def get_product_name(self, two_lang_links, row, item_articule, work_sheet):
        # Get names ru ukr
        for idx, language_link in enumerate(two_lang_links):
            try:
                req = requests.get(language_link, headers=self.headers)
                soup_name = BeautifulSoup(req.text, "lxml")

                manufacturers = ("Noveen ", "Marble")  # Manufacturers names with 1 space after
                series = work_sheet.cell(row, 3).value  # Work file must have 3 columns

                if series is None:
                    series = ""
                else:
                    series = f"{series} "

                # Implement scrapping
                full_name = (soup_name.find("div", id="main")
                             .find("h1").text)

                manufacturer = next((m for m in manufacturers if m in full_name), None)
                item_type, last_name = full_name.split(manufacturer, 1)

                # Diiferent methods to make the right name. Check Readme
                last_name = last_name.replace(f"{series}", "")

                new_name = f"{item_type.strip()} {last_name.strip()} {item_articule} {manufacturer}"
                print(new_name)

                self.blank_sheet.cell(row, 4 + idx).value = new_name

                # Save names data
                self.save_names_data(item_type, last_name, item_articule, series, manufacturer, row, idx)

            except Exception as ex:
                print(ex)
                print("No name or manufacturer name")