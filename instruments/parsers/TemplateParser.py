from bs4 import BeautifulSoup
import os
import requests
import time
import random
from pathlib2 import Path
from urllib.parse import urlparse
import openpyxl

from instruments.BaseParser import BaseParser
from instruments.BaseParser import ParserLogger

# Only for copy and make new data scrapper for new website
class TemplateParser(BaseParser):
    def __init__(self):
        super().__init__()


    def scrap(self, filename: str = "scrapped_data(full).xlsx",
              export_file: str = "export.xlsx"):

        work_sheet = openpyxl.open(export_file).active
        logger = ParserLogger()

        for row in range(2, work_sheet.max_row + 1):
            print(self.GREEN(f"{row}. Started"))

            item_articule = work_sheet.cell(row, 2).value

            url_search_ru = "website searching link"    # Need to add searching link
            url = f"{url_search_ru}{item_articule}"

            searched_item_link = self.get_searched_item_link(url)
            print(searched_item_link)

            self.log_data = [row - 1 ,item_articule, searched_item_link]

            if not searched_item_link:
                print(self.GREEN(f"Link skipped. None response"))
                logger.log_parsing_result(self.log_data)
                continue

            self.blank_sheet.cell(row, 2).value = item_articule

            # If there is necessary to scrap UKR version, it is important for names and descriptions
            ukr_link = searched_item_link.replace(..., ...)  # Old link to new UKR link
            two_lang_links = (searched_item_link, ukr_link)

            req = requests.get(searched_item_link, headers=self.headers)
            soup = BeautifulSoup(req.text, "lxml")

            """Обираємо необхідний парсер і реалізуємо його роботу в окремому файлі"""
            # Uncomment necessary scrappers
            self.get_characteristics(soup, row)
            self.download_instruction(soup, row)
            self.get_descriptions(two_lang_links, row)
            self.get_photos(soup, row)
            self.get_product_name(two_lang_links, row, item_articule, work_sheet)

            logger.log_parsing_result(self.log_data)

            self.blank_file.save(filename)
            time.sleep(1 + random.uniform(1, 3))
        print(self.GREEN(f"\nFile {filename} created"))
        print(self.GREEN(f"Total photo count: {self.images_counter}"))
        print(self.GREEN(f"Total descriptions count: {self.instructions_counter}"))

    def get_searched_item_link(self, url):
        """Шукає товар за артикулом та повертає його URL"""
        try:
            req = requests.get(url, headers=self.headers)
            soup = BeautifulSoup(req.text, "lxml")
            searched_item_link = soup.find(...)  # Реалізація пошуку
            return searched_item_link
        except Exception as ex:
            self.log_data.append(f"Error getting search link: {ex}")
            return None

    def get_characteristics(self, soup, row):
        """Отримує характеристики товару"""
        try:
            characteristics = soup.find(...)  # Get characteristics

            # Get keys values (can be some for find_all loops)
            # Implement scrap

            key, value = ..., ...

            # Check if new char name is not in char dick
            self.check_key(key)

            # Adding characteristics
            char_col = self.char_dict[key]
            self.blank_sheet.cell(row, char_col).value = value

        except Exception as ex:
            self.log_data.append(f"Error getting characteristics: {ex}")

    def download_instruction(self, soup, row):
        # Get instructions
        try:
            # Implement data scrapper
            find_instr = soup.find(...)

            instruction_link = ...

            self.download_instruction_file(instruction_link, row)

        except Exception as ex:
            self.log_data.append(f"Error getting instructions {ex}")

    def get_descriptions(self, two_lang_links, row):
        # Get description RU (But now it is better to generate with gpt)
        # OPTIONAL: Get description UKR if it is possible on website (other option is translator)
        for idx, language_link in enumerate(two_lang_links):
            try:
                req = requests.get(language_link, headers=self.headers)
                soup_descr = BeautifulSoup(req.text, "lxml")

                description_lines = soup_descr.find(...).contents
                clean_descriptions = [str(i) for i in description_lines if i != "\n"]

                self.blank_sheet.cell(row, 11 + idx).value = "\n".join(clean_descriptions)
                time.sleep(1 + random.uniform(1, 2))
            except Exception as ex:
                self.log_data.append(f"Error getting descriptions: {ex}")

    def get_photos(self, soup, row):
        # Get photos
        try:
            # Implement scrapping
            photo_data = (soup.find(...))

            photo_links = [link.get("href") for link in photo_data]

            self.download_photos(photo_links, row, "folder_name") # Input folder name
        except Exception as ex:
            self.log_data.append(f"Error getting photos: {ex}")

    def get_product_name(self, two_lang_links, row, item_articule, work_sheet):
        # Get names ru ukr
        for idx, language_link in enumerate(two_lang_links):
            try:
                req = requests.get(language_link, headers=self.headers)
                soup_name = BeautifulSoup(req.text, "lxml")

                manufacturers = ("Name ", ...)  # Manufacturers names with 1 space after
                series = work_sheet.cell(row, 3).value  # Work file must have 3 columns

                if series is None:
                    series = ""
                else:
                    series = f"{series} "

                # Implement scrapping
                full_name = soup_name.find(...)

                manufacturer = next((m for m in manufacturers if m in full_name), None)
                item_type, last_name = full_name.split(manufacturer, 1)

                # Diiferent methods to make the right name. Check Readme
                last_name = last_name.replace(f"{series}", "")

                new_name = f"{item_type.strip()} {last_name.strip()} {item_articule} {series}{manufacturer}"
                print(new_name)

                self.blank_sheet.cell(row, 4 + idx).value = new_name

                # Save names data
                self.save_names_data("filename", item_type, last_name, item_articule, series, manufacturer, row, idx)

            except Exception as ex:
                self.log_data.append(f"Error getting product name {ex}")