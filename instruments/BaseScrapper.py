import hashlib
import openpyxl
from pathlib2 import Path
from PyPDF2 import PdfReader
from io import BytesIO
from fake_useragent import UserAgent

from instruments.Resources import Resources


class BaseScrapper(Resources):
    def __init__(self):
        super().__init__()

        self.headers = {
            "User-Agent": UserAgent().random,
        }


    # Returns title of the pdf file
    @staticmethod
    def read_pdf(request) -> str:
        pdf_data = BytesIO(request.content)
        pdf_reader = PdfReader(pdf_data)
        metadata = pdf_reader.metadata
        title = metadata.get("/Title", "Назва не знайдена")

        return title

    @staticmethod
    def save_pdf(file_path: Path, request) -> str:
        file_content = request.content  # Отримуємо вміст файлу
        file_stem = file_path.stem  # Початкова назва без розширення
        file_ext = file_path.suffix  # Розширення файлу (.pdf)

        # Генеруємо хеш (беремо 8 символів для унікальності)
        file_hash = hashlib.sha256(file_content).hexdigest()[:8]

        # Формуємо нове ім'я файлу
        new_file_name = f"{file_stem}_{file_hash}{file_ext}"
        new_file_path = file_path.parent / new_file_name

        # Записуємо файл
        with open(new_file_path, "wb") as file:
            file.write(file_content)

        return new_file_name

    def save_names_data(self, item_type, last_name, item_articule, series, manufacturer, row, idx):
        self.names_sheet.cell(row, 1 + idx).value = item_type
        self.names_sheet.cell(row, 3 + idx).value = last_name
        self.names_sheet.cell(row, 5).value = item_articule
        self.names_sheet.cell(row, 6).value = series
        self.names_sheet.cell(row, 7).value = manufacturer

        self.book_names_data.save("Names_data.xlsx")