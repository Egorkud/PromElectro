import time
import openpyxl
from openpyxl import load_workbook
import os
import pandas as pd
from openpyxl.workbook import Workbook
from openpyxl.styles import PatternFill
from pathlib2 import Path
from pdf2image import convert_from_path
from io import BytesIO
import img2pdf
import shutil
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager


from instruments import config
from instruments.Resources import Resources


class DataInstruments(Resources):
    def __init__(self):
        super().__init__()

    def init_project(self):
        def create_path(*files):
            for i in files:
                if not i.exists():
                    if i.suffix:
                        i.touch(exist_ok=True)
                        print(self.GREEN(f"File '{i}' created"))
                    else:
                        i.mkdir(exist_ok=True)
                        print(self.GREEN(f"Directory '{i}' created"))

        def create_excel_file(file_path, columns):
            """Створює Excel-файл з вказаними заголовками, якщо він не існує."""
            if file_path.exists():
                return

            create_path(file_path.parent, file_path)

            wb = Workbook()
            sheet = wb.active
            sheet.title = "Sheet"

            for col_id, col_name in columns.items():
                sheet.cell(1, col_id).value = col_name

            wb.save(file_path)
            print(self.GREEN(f"File {file_path} was filled"))

        print(self.BLUE("\nProject initialisation started\n"))
        # Crete folders (convenience purpose)
        folders = ("import_done", "import_queue", "temp_old", "downloaded_groups")
        create_path(*(Path(i) for i in folders))

        # Create data directory and files inside
        data_dir = Path("data")
        sample_file = data_dir / "sample.xlsx"
        names_data_file = data_dir / "names_data.xlsx"

        create_excel_file(sample_file, config.SAMPLE_PRODUCT_COLUMNS)
        create_excel_file(names_data_file, config.NAMES_DATA_COLUMNS)

        print(self.BLUE("\nProject initialisation finished\n"))

    @staticmethod
    def generate_numbers_string(num_1: int, num_2: int, filename: str = "new_numbers.txt") -> None:
        """
        Creates .txt file with numbers divided by comma
        :param num_1: Start number
        :param num_2: Finish number
        :param filename: Name of file + .txt
        """
        if num_1 > num_2:
            raise ValueError("Цифра 1 більше за цифру 2")

        numbers = [str(i) for i in range(num_1, num_2 + 1)]
        result_string = ",".join(numbers)

        with open(filename, "w", encoding="utf-8") as f:
            f.write(result_string)

    @staticmethod
    def collect_product_numbers(directory: str, output_file: str = "product_numbers.txt") -> None:
        """
        Проходить по всіх Excel-файлах у папці, збирає номери товарів із першої колонки
        та записує їх у текстовий файл через кому без пробілів.

        :param directory: Шлях до папки з Excel файлами
        :param output_file: Ім'я вихідного текстового файлу
        """

        directory = Path(directory)
        if not directory.exists() or not directory.is_dir():
            print(f"❌ Папка {directory} не існує або не є директорією!")
            return

        product_numbers = []

        for idx, file_path in enumerate(directory.glob("*.xls*")):
            try:
                print(f"{idx + 1}. Обробка файлу: {file_path.name}")

                # Визначаємо, який engine використовувати
                engine = "openpyxl" if file_path.suffix == ".xlsx" else "xlrd"
                df = pd.read_excel(file_path, dtype=str, engine=engine)

                if df.shape[1] < 1:
                    print(f"Файл {file_path.name} не містить колонок. Пропускаємо.")
                    continue

                numbers = df.iloc[:, 0].dropna().astype(str).tolist()  # Беремо 1-шу колонку
                product_numbers.extend(numbers)

            except Exception as e:
                print(f"Помилка при обробці {file_path.name}: {e}")

        # Записуємо у файл через кому без пробілів
        with open(output_file, "w", encoding="utf-8") as f:
            f.write(",".join(product_numbers))

        print(f"Готово! Дані записані у {output_file}")


    # Fill descriptions from descriptions sheet.
    # Column 1. Name or id as convenient
    # Column 2. Group name (full path to group).
    def groups_filler(self,
                      filename:str = "new_groups.xlsx",
                      export_file:str = "export.xlsx"):
        groups_dict = {}
        export_file = openpyxl.open(export_file)
        export_sheet = export_file["export sheet"]
        groups_sheet = export_file["groups sheet"]

        for row in range(1, groups_sheet.max_row + 1):
            id_name = groups_sheet.cell(row, 1).value
            group_name = groups_sheet.cell(row, 2).value
            groups_dict.update([(id_name, group_name)])

        for row in range(2, export_sheet.max_row + 1):
            id_name = export_sheet.cell(row, 3).value
            if id_name in groups_dict.keys():
                group_name = groups_dict[id_name]
                export_sheet.cell(row, 3).value = group_name
                print(self.GREEN(f"{row}. changed"))
            else:
                print(self.YELLOW(f"{row}. skipped"))

        export_file.save(filename)
        print(self.GREEN(f"\nFile {filename} created"))


    # Compress all the files by screenshotting pages
    @staticmethod
    def compress_pdf_folder(input_folder:str = "downloaded_pdfs",
                            output_folder:str = "compressed_pdfs",
                            dpi:int = 200):
        # Потрібно завантажити цей інструмент та можна просто додати в
        # директорію проєкту та далі вказати тут шлях до нього
        poppler_path = r"../data/poppler-24.08.0/Library/bin"

        # Створюємо вихідну папку, якщо її немає
        os.makedirs(output_folder, exist_ok=True)

        # Перебираємо всі файли в папці
        for file_name in os.listdir(input_folder):
            if file_name.lower().endswith(".pdf"):
                input_pdf = os.path.join(input_folder, file_name)
                output_pdf = os.path.join(output_folder, file_name)

                images = convert_from_path(input_pdf, dpi=dpi, poppler_path=poppler_path)

                img_bytes = []
                for img in images:
                    img_buffer = BytesIO()
                    img.save(img_buffer, format="JPEG", quality=50)  # Стиснення JPEG
                    img_bytes.append(img_buffer.getvalue())

                with open(output_pdf, "wb") as f:
                    f.write(img2pdf.convert(img_bytes))

                print(f"Стиснуто: {file_name}")


    # Можна поміняти файли місцями, щоб перевірити роботу про всяк випадок
    # Для цього додано параметр reverse_check
    @staticmethod
    def check_duplicates_articule(export_file: str = "name.xlsx",
                                  work_file: str = "name.xlsx",
                                  duplicates_file: str = "duplicates.xlsx",
                                  unique_file: str = "unique.xlsx",
                                  export_cols: tuple = (1, 2),  # (артикул, назва товару)
                                  work_cols: tuple = (1, 2)):

        export_df = pd.read_excel(export_file, usecols=list(export_cols))
        work_df = pd.read_excel(work_file, usecols=list(work_cols))

        wb = load_workbook(work_file, data_only=True)
        ws = wb.active

        # Видаляємо пробіли тільки для порівняння
        export_articles = set(export_df.iloc[:, 0].dropna().astype(str).str.replace(" ", ""))

        work_df = work_df.dropna(subset=[work_df.columns[0]])
        work_df["Артикул_чистий"] = work_df.iloc[:, 0].astype(str).str.replace(" ", "")
        work_df["Артикул"] = work_df.iloc[:, 0].astype(str)
        work_df["Назва"] = work_df.iloc[:, 1]

        work_df["Посилання"] = None
        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=work_cols[1] + 1)
            if cell.hyperlink:
                work_df.at[row_idx - 2, "Посилання"] = cell.hyperlink.target

        work_df["Дублікат"] = work_df["Артикул_чистий"].isin(export_articles)
        duplicates = work_df[work_df["Дублікат"]][["Артикул", "Назва", "Посилання"]].sort_values(by="Артикул")
        unique = work_df[~work_df["Дублікат"]][["Артикул", "Назва", "Посилання"]].sort_values(by="Артикул")

        def save_to_excel(df, filename, sheet_name):
            wb_out = Workbook()
            ws_out = wb_out.active
            ws_out.title = sheet_name
            ws_out.append(["Артикул", "Назва товару", "Посилання"])
            for row in df.itertuples(index=False):
                ws_out.append(list(row))
            wb_out.save(filename)

        save_to_excel(duplicates, duplicates_file, "Дублікати")
        save_to_excel(unique, unique_file, "Унікальні")

        print(f"✅ Дублікати збережено у '{duplicates_file}'")
        print(f"✅ Унікальні товари збережено у '{unique_file}'")


    # Simply gets data from one file and write to empty sheet excel
    def get_coloured_cells(self, file_name:str = "name.xlsx"):
        counter = 1
        work_sheet = openpyxl.open(file_name).active
        for row in range(1, work_sheet.max_row + 1):

            name = work_sheet.cell(row, 1).value
            articule = work_sheet.cell(row, 3).value
            cell_fill = work_sheet.cell(row, 11).fill

            # Checks whether cell is coloured
            if cell_fill.bgColor.rgb != "00000000":
                print(row)
                self.empty_sheet.cell(counter, 1).value = name
                self.empty_sheet.cell(counter, 2).value = articule

                counter += 1

        self.book_empty.save("new_filtered_data.xlsx")

    @staticmethod
    def name_changer(input_file:str = "import_queue/Освещение=Светильники.xlsm",
                     output_file:str = "import_queue/New_names.xlsx",
                     col_name1:int = 4,
                     col_name2:int = 5,
                     col_article:int = 2,
                     col_group:int = 3,
                     selected_group:str = "Освещение=>Светильники=>Бытовые светильники=>Точечные светильники",
                     attribute_cols:set = (109, 100, 77)):
        """
           Оновлює дві колонки назв, додаючи перед артикулом дані з вказаних колонок,
           але тільки для товарів, що належать до заданої групи.
           Рядки, які не підходять, пропускаються.

           :param input_file: Назва вхідного файлу Excel
           :param output_file: Назва вихідного файлу Excel
           :param col_name1: Номер першої колонки з назвами (1-індекс)
           :param col_name2: Номер другої колонки з назвами (1-індекс)
           :param col_article: Номер колонки з артикулами (1-індекс)
           :param col_group: Номер колонки з групою товарів (1-індекс)
           :param selected_group: Назва групи, за якою фільтруємо
           :param attribute_cols: Номери колонок (1-індекс), дані яких вставлятимуться в назву
        """

        # Завантажуємо дані
        df = pd.read_excel(input_file, dtype=str)

        # Конвертуємо індексацію (Excel → Python, тобто віднімаємо 1)
        col_name1 -= 1
        col_name2 -= 1
        col_article -= 1
        col_group -= 1
        attribute_cols = {col - 1 for col in attribute_cols}

        # Отримуємо назви колонок
        name_col1 = df.columns[col_name1]
        name_col2 = df.columns[col_name2]
        article_col = df.columns[col_article]
        group_col = df.columns[col_group]
        attribute_columns = [df.columns[i] for i in attribute_cols]

        def modify_name(row, name_col):
            name = row[name_col]
            article = str(row[article_col])

            # Збираємо атрибути
            attributes = " ".join(str(row[col]) for col in attribute_columns if pd.notna(row[col]))

            # Якщо артикул є у назві — вставляємо перед ним атрибути
            if article in name:
                return name.replace(article, attributes + " " + article)

            return name  # Якщо артикул не знайдено, залишаємо без змін

        # Фільтруємо рядки тільки для обраної групи
        df_filtered = df[df[group_col] == selected_group].copy()

        # Оновлюємо обидві колонки назв
        df_filtered["Оновлена Назва (RU)"] = df_filtered.apply(lambda row: modify_name(row, name_col1), axis=1)
        df_filtered["Оновлена Назва (UA)"] = df_filtered.apply(lambda row: modify_name(row, name_col2), axis=1)

        # Зберігаємо тільки артикул + оновлені назви
        df_result = df_filtered[[df.columns[col_article], "Оновлена Назва (RU)", "Оновлена Назва (UA)"]]

        # Зберігаємо результат без пустих рядків
        df_result.to_excel(output_file, index=False, engine="openpyxl")

        print(f"✅ Файл збережено: {output_file}, збережено {len(df_result)} рядків")

    @staticmethod
    def get_code_row(input_file: str = "website_positions.xlsx",
                     output_file: str = "output_codes.txt"):

        # Зчитуємо тільки 1-й стовпець (index 0) без заголовків
        df = pd.read_excel(input_file, usecols=[0], dtype=str, engine='openpyxl')

        # Видаляємо порожні значення та дублікати
        unique_codes = df.iloc[:, 0].dropna().unique()

        # Швидкий запис у файл через кому
        with open(output_file, "w", encoding="utf-8") as f:
            f.write(",".join(unique_codes))

        print(f"Збережено {len(unique_codes)} кодів у {output_file}")

    @staticmethod
    def extract_unique_categories(folder_path: str = "import_done",
                                  output_file: str = "unique_categories.xlsx",
                                  col_category: int = 3):
        """
            Збирає унікальні категорії товарів з усіх Excel-файлів у папці.

            :param folder_path: Шлях до папки з Excel-файлами.
            :param output_file: Назва файлу для збереження унікальних категорій.
            :param col_category: Номер колонки (1-індекс), у якій знаходиться категорія товару.
            """
        unique_categories = set()

        # Перетворюємо індексацію (Excel → Python, тобто віднімаємо 1)
        col_category -= 1

        # Проходимо по всіх файлах у папці
        for file in Path(folder_path).iterdir():
            if file.suffix in {".xls", ".xlsx", ".xlsm"}:  # Фільтр тільки для Excel-файлів
                try:
                    df = pd.read_excel(file, dtype=str, usecols=[col_category])  # Читаємо лише потрібну колонку
                    unique_categories.update(df.iloc[:, 0].dropna().unique())  # Додаємо унікальні значення
                    print(f"Опрацьовано: {file.name}")
                except Exception as e:
                    print(f"❌ Помилка обробки {file.name}: {e}")

        # Сортуємо категорії
        sorted_categories = sorted(unique_categories)

        # Зберігаємо результат
        df_result = pd.DataFrame(sorted_categories, columns=["Унікальні категорії"])
        df_result.to_excel(output_file, index=False, engine="openpyxl")

        print(f"✅ Файл збережено: {output_file} (всього {len(sorted_categories)} унікальних категорій)")

    @staticmethod
    def download_categories(headless: bool = False,
                            login: str = os.getenv("login"),
                            password: str =  os.getenv("password"),
                            category_file: str = "unique_categories.xlsx",
                            product_range: str = "1-100000"):
        """
            Запускає браузер через Selenium, використовуючи webdriver-manager для автоматичного завантаження драйвера.\n
            Для створення файлу є метод extract_unique_categories().

            :param headless: Якщо True, браузер запускається без інтерфейсу (фоновий режим).
            :param login: Логін потрібно брати зі змінної середовища або передавати методу.
            :param password: Пароль потрібно брати зі змінної середовища або передавати методу.
            :param category_file: Шлях до ексель-файлу з категоріями.
            :param product_range: Діапазон товарів, який потрібно встановити (наприклад, "1-100000").
            """

        if not Path(category_file).exists():
            print(f"Error, there is no such file: {category_file}")
            return

        # region Driver initialisation
        options = webdriver.ChromeOptions()

        if headless:
            options.add_argument("--headless")  # Фоновий режим (без UI)

        options.add_argument("--disable-gpu")
        options.add_argument("--no-sandbox")
        options.add_argument("--start-maximized")  # Відкриває в повному екрані

        # Автоматичне встановлення драйвера
        service = Service(ChromeDriverManager().install())

        # Запускаємо браузер
        driver = webdriver.Chrome(service=service, options=options)
        # endregion

        try:
            # 1. Зчитування категорій з файлу
            df = pd.read_excel(category_file, dtype=str)
            categories = df.iloc[:, 0].dropna().tolist()  # Беремо перший стовпець

            # 2. Вхід на сайт
            driver.get("https://a.electro-market.com.ua/")
            WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//input[@type="text"]')))

            # Вводимо логін і пароль та автентифікуємося
            driver.find_element(By.XPATH, '//input[@type="text"]').send_keys(login)
            driver.find_element(By.XPATH, '//input[@type="password"]').send_keys(password + Keys.RETURN)
            WebDriverWait(driver, 10).until(EC.url_contains("s_admin"))  # Чекаємо на редірект

            print("Авторизація успішна!")

            # 3. Перехід на сторінку експорту
            driver.get("https://a.electro-market.com.ua/s_admin/ru/catalogue/import-export/export/")
            WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, "//button[contains(text(), 'Загрузить')]")))

            print("Перехід до сторінки експорту...")

            # 4. Зміна значення "1-5000" на власне
            product_range_input = driver.find_element(By.ID, "products")
            product_range_input.clear()  # Очистка поля
            product_range_input.send_keys(product_range)
            print(f"Items range: {product_range}")
            time.sleep(1)

            print(f"Знайдено {len(categories)} категорій!")

            for idx, category in enumerate(categories):
                print(f"\n🔎 {idx + 1}. Обробка категорії: {category}")

                levels = category.split("=>")  # Розділяємо рівні вкладеності
                last_level = levels[-1].strip()  # Беремо лише останній рівень категорії

                try:
                    # 5. Знаходимо та клікаємо лише на останню категорію
                    last_xpath = f"//span[contains(text(), '{last_level}')]"
                    category_element = WebDriverWait(driver, 10).until(
                        EC.element_to_be_clickable((By.XPATH, last_xpath)))

                    # Використовуємо ActionChains для точного кліку
                    ActionChains(driver).move_to_element(category_element).click().perform()
                    time.sleep(1)

                    print(f"✅ Категорія '{last_level}' вибрана!")

                    # 6. Натискаємо кнопку "Загрузить"
                    download_button = driver.find_element(By.XPATH, "//button[contains(text(), 'Загрузить')]")
                    download_button.click()
                    print(f"📥 Завантаження для '{last_level}' розпочато!")
                    time.sleep(8)  # Чекаємо завантаження, якщо виникатимуть помилки - збільшити час

                    # 7. Повторне натискання для скидання попереднього вибору
                    ActionChains(driver).move_to_element(category_element).click().perform()
                    time.sleep(1)

                except Exception as e:
                    print(f"❌ Не вдалося знайти категорію '{category}': {e}")

            print("\n✅ Завантаження завершено!")

        finally:
            driver.quit()  # Закриваємо браузер після завершення роботи

    @staticmethod
    def process_excel_files(directory: str = "downloaded_groups",
                            article_file: str = None,
                            new_data_file: str = None,
                            max_columns: int = 42):
        """
        Обробляє всі Excel-файли у вказаній папці:
        - Перевіряє, чи всі товари мають одну категорію у 3-й колонці, та перейменовує файл.
        - Конвертує .xls → .xlsx зі збереженням форматування.
        - Форматує Excel (row height 15, freeze top row, auto filter).
        - Зафарбовує рядки у блакитний колір, якщо артикул є в наданому файлі.
        - Додає нові дані з файлу new_data_file у відповідні категорії.

        :param directory: Шлях до папки з Excel файлами.
        :param article_file: Шлях до файлу з артикулами (опціонально).
        :param new_data_file: Шлях до файлу з новими даними (опціонально).
        :param max_columns: Кількість нових доданих колонок у файл
        """

        def sanitize_filename(name: str) -> str:
            invalid_chars = ('<', '>', ':', '"', '/', '\\', '|', '?', '*')
            for char in invalid_chars:
                name = name.replace(char, '')
            return name.strip()

        def detect_excel_format(file_path: Path):
            """ Визначає формат Excel-файлу, навіть якщо розширення неправильне """
            try:
                with open(file_path, "rb") as f:
                    header = f.read(8)
                if header.startswith(b"\xD0\xCF\x11\xE0"):  # Старий .xls (OLE2)
                    return "xls"
                elif header.startswith(b"PK\x03\x04"):  # Новий .xlsx (ZIP-based)
                    return "xlsx"
                else:
                    return None
            except Exception:
                return None

        def format_excel(file_path: Path, highlight_articles=set(), new_rows=set()):
            """ Форматує Excel-файл та фарбує рядки """
            try:
                wb = openpyxl.load_workbook(file_path)
                ws = wb.active

                # Встановлення висоти рядків
                for row in ws.iter_rows():
                    ws.row_dimensions[row[0].row].height = 15

                # Закріплення заголовка
                ws.freeze_panes = "A2"

                # Увімкнення автофільтра
                ws.auto_filter.ref = ws.dimensions

                # Фарбування рядків у блакитний, якщо артикул у списку
                blue_fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
                for row in ws.iter_rows(min_row=2):  # Пропускаємо заголовки
                    if row[1].value in highlight_articles or row[0].row in new_rows:
                        for cell in row:
                            cell.fill = blue_fill

                wb.save(file_path)
                print(f"✅ Файл відформатовано: {file_path.name}")
            except Exception as e:
                print(f"⚠️ Помилка при форматуванні {file_path.name}: {e}")

        directory = Path(directory)
        if not directory.exists():
            print(f"❌ Папка {directory} не існує!")
            return

        # Зчитуємо артикул-файл, якщо передано
        highlight_articles = set()
        if article_file and Path(article_file).exists():
            try:
                df_articles = pd.read_excel(article_file, dtype=str, engine="openpyxl")
                highlight_articles = set(df_articles.iloc[:, 1].dropna().unique())  # Колонка 2 (індекс 1)
                print(f"🔹 Зчитано {len(highlight_articles)} унікальних артикулів для виділення")
            except Exception as e:
                print(f"⚠️ Помилка при зчитуванні файлу з артикулами: {e}")

        # Зчитуємо нові дані, якщо передано
        new_data = {}
        if new_data_file and Path(new_data_file).exists():
            try:
                df_new = pd.read_excel(new_data_file, dtype=str, engine="openpyxl")
                if df_new.shape[1] >= 3:
                    for _, row in df_new.iterrows():
                        category = sanitize_filename(row.iloc[2])  # 3-я колонка – категорія
                        if category and category not in new_data:
                            new_data[category] = []
                        new_data[category].append(row.tolist())  # Додаємо новий рядок
                    print(f"🔹 Нові дані розподілено по категоріях: {len(new_data)} категорій")
            except Exception as e:
                print(f"⚠️ Помилка при зчитуванні файлу з новими даними: {e}")

        for idx, file_path in enumerate(directory.glob("*.xls*")):
            try:
                print(f"\n📂 {idx + 1}. Обробка файлу: {file_path.name}")

                # Визначаємо формат
                detected_format = detect_excel_format(file_path)
                if detected_format == "xls":
                    df = pd.read_excel(file_path, dtype=str, engine="xlrd")  # Старий .xls
                elif detected_format == "xlsx":
                    df = pd.read_excel(file_path, dtype=str, engine="openpyxl")  # Новий .xlsx
                else:
                    print(f"⚠️ Файл {file_path.name} не є дійсним Excel. Пропускаємо.")
                    continue

                if df.shape[1] < 3:
                    print(f"⚠️ Файл {file_path.name} містить менше 3 колонок. Пропускаємо.")
                    continue

                categories = df.iloc[:, 2].dropna().unique()

                if len(categories) == 1:
                    category_name = sanitize_filename(categories[0])
                    new_filename = category_name + ".xlsx"
                    new_path = directory / new_filename

                    if new_path.exists():
                        print(f"⚠️ Файл {new_filename} вже існує. Пропускаємо.")
                        continue

                    # Перетворення .xls у .xlsx
                    if detected_format == "xls":
                        temp_xlsx = file_path.with_suffix(".xlsx")
                        df.to_excel(temp_xlsx, index=False, engine="openpyxl")
                        shutil.copy2(temp_xlsx, new_path)
                        temp_xlsx.unlink()
                    else:
                        shutil.copy2(file_path, new_path)

                    file_path.unlink()

                    # Додаємо нові дані до файлу
                    if new_data_file and category_name in new_data:
                        wb = openpyxl.load_workbook(new_path)
                        ws = wb.active
                        start_row = ws.max_row + 1

                        for row_idx, new_row in enumerate(new_data[category_name], start=start_row):
                            formatted_row = new_row[:max_columns] + [""] * (max_columns - len(new_row))
                            ws.append(formatted_row)  # Додаємо новий рядок

                            for cell in ws[row_idx]:
                                cell.fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5",
                                                        fill_type="solid")  # Заливка
                        wb.save(new_path)
                        print(f"✅ Додано {len(new_data[category_name])} нових рядків у {new_filename}")

                        format_excel(new_path, highlight_articles, new_rows=set(range(start_row, ws.max_row + 1)))
                    else:
                        format_excel(new_path, highlight_articles)

                    print(f"✅ Файл перейменовано: {file_path.name} -> {new_filename}")
                else:
                    print(f"⚠️ У файлі {file_path.name} кілька різних категорій, не перейменовується.")

            except Exception as e:
                print(f"❌ Помилка при обробці файлу {file_path.name}: {e}")

    @staticmethod
    def merge_xlsx_files(input_folder: str = "import_queue",
                         output_file: str = "new_merged_data.xlsx") -> None:
        """
        Merges all data from .xlsx files in input_folder.

        :param input_folder: Folder where .xlsx files are located.
        :param output_file: Name of output file.
        """
        # Створюємо шлях до папки з файлами
        folder_path = Path(input_folder)

        # Перевіряємо, чи існує така папка
        if not folder_path.exists() or not folder_path.is_dir():
            raise ValueError(f"Папка {input_folder} не існує або це не директорія")

        # Змінна для зберігання даних усіх файлів
        merged_data = []

        # Проходимо по кожному xlsx файлу в папці
        for file in folder_path.glob("*.xlsx"):
            # Завантажуємо файл
            df = pd.read_excel(file)

            # Додаємо ці дані в список
            merged_data.append(df)

        # Об'єднуємо всі дані в один DataFrame
        merged_df = pd.concat(merged_data, ignore_index=True)

        # Записуємо результат в новий xlsx файл
        merged_df.to_excel(output_file, index=False)

        print(f"Дані успішно збережено в {output_file}")