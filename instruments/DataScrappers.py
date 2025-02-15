from bs4 import BeautifulSoup
import os
import requests
import time
import random
import openpyxl
from pathlib2 import Path
from urllib.parse import urlparse
from fake_useragent import UserAgent


from instruments.Resources import Resources



class DataScrappers(Resources):
    def __init__(self):
        super().__init__()

        self.headers = {
            "User-Agent": UserAgent().random,
        }

    def scrap(self, filename : str = "scrapped_data (instructions+photos).xlsx",
              export_file : str = "export.xlsx"):
        char_dict = {}
        counter = 43
        work_sheet = openpyxl.open(export_file).active

        for row in range(2, work_sheet.max_row + 1):
            print(self.GREEN(f"{row}. Started"))

            item_articule = work_sheet.cell(row, 2).value
            self.blank_sheet.cell(row, 2).value = item_articule

            url_search_ru = "website searching ling"    # Need to add searching link
            url = f"{url_search_ru}{item_articule}"

            searched_item_link = self.get_searched_item_link(url)
            print(searched_item_link)

            if not searched_item_link:
                print(self.GREEN(f"Link skipped. None response"))
                continue

            # If there is necessary to scrap UKR version, it is important for names and descriptions
            ukr_link = searched_item_link.replace(..., ...)  # Old link to new UKR link
            two_lang_links = [searched_item_link, ukr_link]

            req = requests.get(searched_item_link, headers=self.headers)
            soup = BeautifulSoup(req.text, "lxml")

            """Обираємо необхідний парсер і реалізуємо його роботу в своїй гілці"""
            # Uncomment necessary scrappers
            self.get_characteristics(soup, char_dict, counter, row)
            self.download_instruction(soup, row)
            self.get_descriptions(two_lang_links, row)
            self.get_photos(soup, row)
            self.get_product_name(two_lang_links, row, item_articule, work_sheet)


            self.blank_file.save(filename)
            time.sleep(1 + random.uniform(1, 3))
        print(self.GREEN(f"\nFile {filename} created"))

    def get_searched_item_link(self, url):
        """Шукає товар за артикулом та повертає його URL"""
        try:
            req = requests.get(url, headers=self.headers)
            soup = BeautifulSoup(req.text, "lxml")
            searched_item_link = soup.find(...)  # Реалізація пошуку
            return searched_item_link
        except Exception as ex:
            print(self.RED(f"Error getting search link: {ex}"))
            return None

    def get_characteristics(self, soup, char_dict, counter, row):
        """Отримує характеристики товару"""
        try:
            characteristics = soup.find(...)  # Get characteristics

            # Get keys values (can be some for find_all loops)
            # Implement scrap

            key, value = ..., ...

            # Check if new char name is not in char dick
            if key not in char_dict.keys():
                char_dict.update([(key, counter)])
                self.blank_sheet.cell(1, counter).value = key
                counter += 1

            # Adding characteristics
            char_col = char_dict[key]
            self.blank_sheet.cell(row, char_col).value = value

        except Exception as ex:
            print(ex)
            print("No characteristics")

    def download_instruction(self, soup, row):
        # Get instructions
        try:
            # Implement data scrapper
            find_instr = soup.find(...)

            instruction_link = ...

            output_folder = "downloaded_pdfs"
            os.makedirs(output_folder, exist_ok=True)

            time.sleep(1 + random.uniform(1, 2))
            req_pdf = requests.get(instruction_link, headers=self.headers)
            title = self.read_pdf(req_pdf)
            file_name = Path(title).stem
            file_path_no_hash = Path(output_folder) / file_name

            file_name_with_hash = self.save_pdf(file_path_no_hash, req_pdf)

            server_file_path = f"/content/instructions/{file_name_with_hash}"
            self.blank_sheet.cell(row, 7).value = server_file_path

        except Exception as ex:
            print(ex)
            print("No instructions")

    def get_descriptions(self, two_lang_links, row):
        # Get description RU (But now it is better to generate with gpt)
        # OPTIONAL: Get description UKR if it is possible on website (other option is translator)
        for idx, language_link in enumerate(two_lang_links):
            try:
                req = requests.get(language_link, headers=self.headers)
                soup_descr = BeautifulSoup(req.text, "lxml")

                description_lines = soup_descr.find(...).contents
                clean_descriptions = [str(i) for i in description_lines if i != "\n"]

                self.blank_sheet.cell(row, 11 + idx).value = "\n".join(clean_descriptions)
                time.sleep(1 + random.uniform(1, 2))
            except Exception as ex:
                print(ex)
                print("No description")

    def get_photos(self, soup, row):
        # Get photos
        try:
            # Implement scrapping
            photo_data = (soup.find(...))

            photo_links = [link.get("href") for link in photo_data]

            for id, link in enumerate(photo_links):
                try:
                    output_folder = "downloaded_photos"
                    os.makedirs(output_folder, exist_ok=True)

                    file_path_name = os.path.basename(urlparse(link).path)
                    file_path_no_hash = os.path.join(output_folder, file_path_name)
                    photo_path_name = f"/content/images/ctproduct_image/FOLDER_NAME"  # Needs INTPUT
                    file_name = f"{file_path_name}_{id + 1}.jpg"

                    time.sleep(0.5 + random.uniform(1, 2))
                    req = requests.get(link, headers=self.headers)
                    with open(f"{file_path_no_hash}_{id + 1}.jpg", "wb") as file:
                        file.write(req.content)

                    self.blank_sheet.cell(row, 16 + id).value = f"{photo_path_name}/{file_name}"
                except:
                    pass

        except Exception as ex:
            print(ex)
            print("No photos")

    def get_product_name(self, two_lang_links, row, item_articule, work_sheet):
        # Get names ru ukr
        for idx, language_link in enumerate(two_lang_links):
            try:
                req = requests.get(language_link, headers=self.headers)
                soup_name = BeautifulSoup(req.text, "lxml")

                manufacturers = ("Name ", ...)  # Manufacturers names with 1 space after
                series = work_sheet.cell(row, 3).value  # Work file must have 3 columns

                if series is None:
                    series = ""
                else:
                    series = f"{series} "

                # Implement scrapping
                full_name = soup_name.find(...)

                manufacturer = next((m for m in manufacturers if m in full_name), None)
                item_type, last_name = full_name.split(manufacturer, 1)

                # Diiferent methods to make the right name. Check Readme
                last_name = last_name.replace(f"{series}", "")

                new_name = f"{item_type.strip()} {last_name.strip()} {item_articule} {series}{manufacturer}"
                print(new_name)

                self.blank_sheet.cell(row, 4 + idx).value = new_name

            except Exception as ex:
                print(ex)
                print("No name or manufacturer name")